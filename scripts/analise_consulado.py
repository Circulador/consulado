# -*- coding: utf-8 -*-
"""
Engenharia reversa do processo interno do Consulado Italiano (RJ)
a partir de 3 planilhas colaborativas (B, P, JSI).
Sistema parcialmente observavel — dados censurados.
"""
import os, re, json, math, datetime as dt
import numpy as np
import openpyxl
from scipy import stats
from sklearn.mixture import GaussianMixture
from sklearn.cluster import KMeans, DBSCAN

np.random.seed(42)
TODAY = dt.date(2026, 7, 17)   # "Atualizada em" nas 3 planilhas

FILES = {
 'B':  r'C:\Users\81008280\Downloads\B APOS ENTREGA NO BALCAO ATE EMAIL  JA SOU ITALIANO (10).xlsx',
 'JSI':r'C:\Users\81008280\Downloads\JSI apos receber e mail Ja sou Italiano aguardando finalizaco consular RJ (1).xlsx',
 'P':  r'C:\Users\81008280\Downloads\P Espera por Pendencias Tempo Consulado SEDEX (1).xlsx',
}
# layout por planilha: (col_entrega, col_nome, col_grupo, col_resol)
LAYOUT = {'B':(0,1,2,4),'JSI':(0,1,2,4),'P':(1,0,2,4)}

def to_date(v):
    if isinstance(v,(dt.datetime,)): return v.date()
    if isinstance(v,(dt.date,)): return v
    if isinstance(v,str):
        m=re.match(r'(\d{4})-(\d{2})-(\d{2})',v)
        if m: return dt.date(int(m[1]),int(m[2]),int(m[3]))
        m=re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})',v)
        if m:
            y=int(m[3]); y=y+2000 if y<100 else y
            return dt.date(y,int(m[2]),int(m[1]))
    return None

BAD = re.compile(r'^(MEDIA|M[EÉ]DIA|MENOR|MAIOR|TOTAL|ULTIMO|[UÚ]LTIMO|ESTAMOS|DATA DE|NOME|RECEBIDO|GRUPO|N[ºo°]|B PLAN|PLANILHA|ATUALIZAD|DIAS|NAN)',re.I)

def load(pid):
    ce,cn,cg,cr = LAYOUT[pid]
    wb = openpyxl.load_workbook(FILES[pid], data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    recs=[]
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i<2: continue
        if row is None or len(row)<=max(ce,cn,cr): continue
        nome = row[cn]
        nome = str(nome).strip() if nome is not None else ''
        ent  = to_date(row[ce])
        if not nome or BAD.match(nome) or ent is None: continue
        res  = to_date(row[cr])
        grp  = row[cg]; grp = str(grp).strip() if grp is not None else '-'
        recs.append(dict(planilha=pid,nome=re.sub(r'\s+',' ',nome),grupo=grp,
                         entrega=ent,resol=res,
                         done=res is not None,
                         dur=(res-ent).days if res else (TODAY-ent).days))
    return recs

DATA={pid:load(pid) for pid in FILES}
out={}

# ---------------------------------------------------------------
# 1. QUALIDADE DOS DADOS
# ---------------------------------------------------------------
def quality(pid):
    r=DATA[pid]; n=len(r)
    neg=[x for x in r if x['dur']<0]
    dup={}
    for x in r:
        k=(x['nome'].upper(),x['entrega'])
        dup[k]=dup.get(k,0)+1
    exact_dup=sum(v-1 for v in dup.values() if v>1)
    # homonimos: mesmo nome, entregas diferentes
    byname={}
    for x in r: byname.setdefault(x['nome'].upper(),set()).add(x['entrega'])
    homon=sum(1 for k,v in byname.items() if len(v)>1)
    fut=[x for x in r if x['resol'] and x['resol']>TODAY]
    return dict(n=n,done=sum(x['done'] for x in r),cens=sum(not x['done'] for x in r),
                neg_dur=len(neg),exact_dup=exact_dup,homonimos=homon,resol_futuro=len(fut),
                pct_censura=round(100*sum(not x['done'] for x in r)/n,1))
out['qualidade']={pid:quality(pid) for pid in DATA}

# sanitiza: remove durações negativas para análise de tempo
for pid in DATA:
    DATA[pid]=[x for x in DATA[pid] if x['dur']>=0]

# ---------------------------------------------------------------
# 2. DESCRITIVAS (somente concluidos p/ duração observada)
# ---------------------------------------------------------------
def desc(pid):
    d=np.array([x['dur'] for x in DATA[pid] if x['done']],float)
    c=np.array([x['dur'] for x in DATA[pid] if not x['done']],float)
    if len(d)==0: return {}
    q=lambda p:float(np.percentile(d,p))
    return dict(n_done=len(d),n_cens=len(c),
        media=round(float(d.mean()),1),mediana=round(float(np.median(d)),1),
        std=round(float(d.std(ddof=1)),1) if len(d)>1 else None,
        var=round(float(d.var(ddof=1)),1) if len(d)>1 else None,
        cv=round(float(d.std(ddof=1)/d.mean()),2) if len(d)>1 else None,
        min=float(d.min()),max=float(d.max()),
        p10=q(10),p25=q(25),p50=q(50),p75=q(75),p90=q(90),p95=q(95),iqr=round(q(75)-q(25),1),
        skew=round(float(stats.skew(d)),2),kurt=round(float(stats.kurtosis(d)),2),
        cens_media=round(float(c.mean()),1) if len(c) else None,
        cens_max=float(c.max()) if len(c) else None)
out['descritivas']={pid:desc(pid) for pid in DATA}

# ---------------------------------------------------------------
# 3. DISTRIBUIÇÕES: normalidade, lognormal, multimodalidade (BIC GMM)
# ---------------------------------------------------------------
def dist_analysis(pid):
    d=np.array([x['dur'] for x in DATA[pid] if x['done'] and x['dur']>0],float)
    res={'n':len(d)}
    if len(d)<8: return res
    # normalidade
    res['shapiro_p']=round(float(stats.shapiro(d).pvalue),4)
    ld=np.log(d)
    res['shapiro_log_p']=round(float(stats.shapiro(ld).pvalue),4)
    # GMM BIC 1..4 componentes em log-duração
    X=ld.reshape(-1,1)
    bic={}
    best=None;bestbic=1e18
    for k in range(1,5):
        if k>len(d): break
        g=GaussianMixture(k,covariance_type='full',n_init=5,random_state=0).fit(X)
        b=g.bic(X); bic[k]=round(float(b),1)
        if b<bestbic: bestbic=b; best=k
    res['bic']=bic; res['k_otimo']=best
    if best and best>1:
        g=GaussianMixture(best,covariance_type='full',n_init=5,random_state=0).fit(X)
        comps=[]
        for i in range(best):
            comps.append(dict(peso=round(float(g.weights_[i]),2),
                              centro_dias=round(float(np.exp(g.means_[i,0])),0),
                              sigma_log=round(float(np.sqrt(g.covariances_[i,0,0])),2)))
        comps.sort(key=lambda c:c['centro_dias'])
        res['componentes']=comps
    # dip-like: bimodalidade via razao BIC
    return res
out['distribuicoes']={pid:dist_analysis(pid) for pid in DATA}

# ---------------------------------------------------------------
# 4. CLUSTERIZAÇÃO (KMeans / GMM / DBSCAN) em concluidos
# ---------------------------------------------------------------
def clustering(pid):
    recs=[x for x in DATA[pid] if x['done'] and x['dur']>0]
    if len(recs)<12: return {'n':len(recs),'nota':'amostra insuficiente'}
    d=np.array([x['dur'] for x in recs],float)
    X=np.log(d).reshape(-1,1)
    Xs=(X-X.mean())/X.std()
    resu={'n':len(recs)}
    # KMeans silhouette 2..4
    from sklearn.metrics import silhouette_score
    sils={}
    for k in range(2,5):
        km=KMeans(k,n_init=10,random_state=0).fit(Xs)
        if len(set(km.labels_))>1:
            sils[k]=round(float(silhouette_score(Xs,km.labels_)),3)
    resu['kmeans_silhouette']=sils
    # DBSCAN
    db=DBSCAN(eps=0.5,min_samples=4).fit(Xs)
    lab=db.labels_
    resu['dbscan_clusters']=int(len(set(lab))-(1 if -1 in lab else 0))
    resu['dbscan_ruido']=int((lab==-1).sum())
    return resu
out['clusterizacao']={pid:clustering(pid) for pid in DATA}

# ---------------------------------------------------------------
# 5 & 6. SURVIVAL (Kaplan-Meier) + HAZARD (life-table)
# ---------------------------------------------------------------
def kaplan_meier(recs):
    # recs: list of (dur, event) event=1 done
    times=sorted(set(x['dur'] for x in recs))
    n=len(recs)
    surv=[]; S=1.0; at_risk=n; var_sum=0.0
    data=sorted(recs,key=lambda x:x['dur'])
    idx=0
    arr=[(x['dur'],1 if x['done'] else 0) for x in data]
    for t in times:
        d_i=sum(1 for (tt,e) in arr if tt==t and e==1)   # eventos
        c_i=sum(1 for (tt,e) in arr if tt==t)            # saídas totais (evento+censura) nesse t
        r_i=sum(1 for (tt,e) in arr if tt>=t)            # em risco
        if r_i==0: continue
        if d_i>0:
            S*=(1-d_i/r_i)
            var_sum+=d_i/(r_i*(r_i-d_i)) if r_i>d_i else 0
        se=S*math.sqrt(var_sum) if var_sum>0 else 0
        surv.append((t,round(S,4),d_i,r_i,round(se,4)))
    # mediana de sobrevivência
    med=None
    for (t,s,_,_,_) in surv:
        if s<=0.5: med=t; break
    return surv,med

def hazard_lifetable(recs,bin_w=30,tmax=None):
    durs=[x['dur'] for x in recs]
    if not durs: return []
    tmax=tmax or (max(durs)+bin_w)
    bins=list(range(0,int(tmax)+bin_w,bin_w))
    rows=[]
    arr=[(x['dur'],1 if x['done'] else 0) for x in recs]
    for i in range(len(bins)-1):
        lo,hi=bins[i],bins[i+1]
        r_i=sum(1 for (t,e) in arr if t>=lo)          # entram no intervalo
        d_i=sum(1 for (t,e) in arr if lo<=t<hi and e==1)  # eventos no intervalo
        w_i=sum(1 for (t,e) in arr if lo<=t<hi and e==0)  # censurados no intervalo
        eff=r_i-w_i/2
        h=d_i/eff if eff>0 else 0
        rows.append(dict(ini=lo,fim=hi,risco=r_i,eventos=d_i,cens=w_i,
                         hazard=round(h,3)))
    return rows

surv_out={}; haz_out={}
for pid in DATA:
    recs=DATA[pid]
    km,med=kaplan_meier(recs)
    # amostra a curva em pontos-chave
    pts=[30,60,90,120,150,180,240,300,365,450,547,730]
    curve={}
    for p in pts:
        s=[x[1] for x in km if x[0]<=p]
        curve[p]=s[-1] if s else 1.0
    surv_out[pid]=dict(mediana_sobrevivencia=med,S_em=curve,n=len(recs),
                       eventos=sum(x['done'] for x in recs))
    haz_out[pid]=hazard_lifetable(recs,bin_w=30)
out['survival']=surv_out
out['hazard']=haz_out

# ---------------------------------------------------------------
# 8. QUEBRA DE FIFO (inversões entre concluídos)
# ---------------------------------------------------------------
def fifo(pid):
    r=[x for x in DATA[pid] if x['done']]
    r=sorted(r,key=lambda x:x['entrega'])
    n=len(r); inv=0; pairs=n*(n-1)//2
    examples=[]
    for i in range(n):
        for j in range(i+1,n):
            # i entrou antes de j (entrega). Inversão se j resolvido antes de i
            if r[i]['entrega']<r[j]['entrega'] and r[j]['resol']<r[i]['resol']:
                inv+=1
                if len(examples)<8 and (r[i]['entrega']-r[j]['entrega']).days<-20:
                    examples.append(dict(cedo=r[i]['nome'],cedo_ent=str(r[i]['entrega']),cedo_res=str(r[i]['resol']),
                                         tarde=r[j]['nome'],tarde_ent=str(r[j]['entrega']),tarde_res=str(r[j]['resol']),
                                         gap_entrada=(r[j]['entrega']-r[i]['entrega']).days,
                                         gap_saida=(r[i]['resol']-r[j]['resol']).days))
    # overtaking global: aguardando ha muito enquanto posteriores sairam
    overtaken=0
    waiting=[x for x in DATA[pid] if not x['done']]
    done=[x for x in DATA[pid] if x['done']]
    for w in waiting:
        later_done=sum(1 for d in done if d['entrega']>w['entrega'])
        if later_done>0: overtaken+=1
    return dict(pares=pairs,inversoes=inv,pct_inversao=round(100*inv/pairs,1) if pairs else 0,
                aguardando_ultrapassados=overtaken,total_aguardando=len(waiting),exemplos=examples)
out['fifo']={pid:fifo(pid) for pid in DATA}

# ---------------------------------------------------------------
# 9. PROCESSAMENTO EM LOTE (dias com muitas resoluções)
# ---------------------------------------------------------------
def batch(pid):
    r=[x for x in DATA[pid] if x['done']]
    by={}
    for x in r: by[x['resol']]=by.get(x['resol'],0)+1
    n=len(r); days=len(by)
    if days==0: return {}
    lam=n/days   # média por dia com evento
    big=sorted(((str(k),v) for k,v in by.items() if v>=3),key=lambda t:-t[1])[:10]
    # teste: prob Poisson de ver >=maxcount
    maxc=max(by.values())
    p_pois=float(1-stats.poisson.cdf(maxc-1,lam))
    # concentração: % de eventos nos top 10% dias mais movimentados
    vals=sorted(by.values(),reverse=True)
    top=vals[:max(1,days//10)]
    conc=round(100*sum(top)/n,1)
    return dict(dias_com_evento=days,media_por_dia=round(lam,2),max_em_um_dia=maxc,
                p_poisson_max=round(p_pois,5),lotes_grandes=big,
                concentracao_top10pct=conc)
out['lote']={pid:batch(pid) for pid in DATA}

# ---------------------------------------------------------------
# 10. MUDANÇA DE REGIME (throughput mensal + change-point)
# ---------------------------------------------------------------
def regime(pid):
    r=[x for x in DATA[pid] if x['done']]
    by={}
    for x in r:
        k=(x['resol'].year,x['resol'].month)
        by.setdefault(k,[]).append(x['dur'])
    months=sorted(by)
    series=[(f'{y}-{m:02d}',len(by[(y,m)]),round(float(np.median(by[(y,m)])),0)) for (y,m) in months]
    # change-point simples: split que maximiza |t| na contagem mensal
    counts=np.array([s[1] for s in series],float)
    cp=None
    if len(counts)>=6:
        best_t=0;best_i=None
        for i in range(2,len(counts)-2):
            a,b=counts[:i],counts[i:]
            if len(a)>1 and len(b)>1 and (a.std()+b.std())>0:
                t=abs(a.mean()-b.mean())/math.sqrt(a.var(ddof=1)/len(a)+b.var(ddof=1)/len(b)+1e-9)
                if t>best_t: best_t=t;best_i=i
        if best_i is not None:
            cp=dict(mes_ruptura=series[best_i][0],t_stat=round(best_t,2),
                    media_antes=round(float(counts[:best_i].mean()),1),
                    media_depois=round(float(counts[best_i:].mean()),1))
    return dict(serie_mensal=series,change_point=cp)
out['regime']={pid:regime(pid) for pid in DATA}

# ---------------------------------------------------------------
# 11. HIPÓTESE PASTA FAMILIAR (proxy) + Bayes + regimes
# ---------------------------------------------------------------
def family_proxy(pid):
    r=[x for x in DATA[pid] if x['done'] and x['dur']>0]
    # suffix compartilhado (padrão FILHO-DANTE) => família
    suffix={}
    for x in r:
        parts=re.split(r'[-–/]',x['nome'])
        if len(parts)>=2:
            suf=parts[-1].strip().upper()
            if len(suf)>=2: suffix.setdefault(suf,[]).append(x)
    shared_suf={k:v for k,v in suffix.items() if len(v)>1}
    fam_ids=set()
    for k,v in shared_suf.items():
        for x in v: fam_ids.add(id(x))
    for x in r:
        if re.search(r'IRM[ÃA]O?S?|IRMA',x['nome'],re.I): fam_ids.add(id(x))
    fam=[x for x in r if id(x) in fam_ids]
    solo=[x for x in r if id(x) not in fam_ids]
    if len(fam)<5 or len(solo)<5:
        return dict(n_familia=len(fam),n_solo=len(solo),nota='amostra insuficiente p/ teste')
    df=np.array([x['dur'] for x in fam],float)
    ds=np.array([x['dur'] for x in solo],float)
    u=stats.mannwhitneyu(df,ds,alternative='two-sided')
    return dict(n_familia=len(fam),n_solo=len(solo),
                mediana_familia=round(float(np.median(df)),0),
                mediana_solo=round(float(np.median(ds)),0),
                media_familia=round(float(df.mean()),0),media_solo=round(float(ds.mean()),0),
                mannwhitney_p=round(float(u.pvalue),4),
                familia_mais_rapida=bool(np.median(df)<np.median(ds)),
                grupos_familiares=len(shared_suf))
out['pasta_familiar']={pid:family_proxy(pid) for pid in DATA}

# Bayes: P(regime rápido | duração) usando GMM 2 comp em JSI (maior amostra)
def bayes_regime(pid):
    d=np.array([x['dur'] for x in DATA[pid] if x['done'] and x['dur']>0],float)
    if len(d)<20: return {}
    X=np.log(d).reshape(-1,1)
    g=GaussianMixture(2,n_init=5,random_state=0).fit(X)
    order=np.argsort(g.means_.ravel())
    fast,slow=order[0],order[1]
    tests=[45,70,100,150,220,310]
    res={}
    for t in tests:
        p=g.predict_proba(np.log([[t]]))[0]
        res[t]=round(float(p[fast]),2)
    return dict(centro_rapido=round(float(np.exp(g.means_[fast,0])),0),
                centro_lento=round(float(np.exp(g.means_[slow,0])),0),
                peso_rapido=round(float(g.weights_[fast]),2),
                P_rapido_dado_dias=res)
out['bayes']={pid:bayes_regime(pid) for pid in DATA}

# ---------------------------------------------------------------
# 12. OUTLIERS explicados
# ---------------------------------------------------------------
def outliers(pid):
    r=[x for x in DATA[pid] if x['done']]
    d=np.array([x['dur'] for x in r],float)
    if len(d)<8: return {}
    q1,q3=np.percentile(d,[25,75]); iqr=q3-q1
    hi=q3+1.5*iqr
    outs=sorted([x for x in r if x['dur']>hi],key=lambda x:-x['dur'])[:10]
    # aguardando extremos
    wait=sorted([x for x in DATA[pid] if not x['done']],key=lambda x:-x['dur'])[:6]
    ex=[]
    for x in outs:
        later_done=sum(1 for y in r if y['entrega']>x['entrega'] and y['resol']<x['resol'])
        ex.append(dict(nome=x['nome'],dias=x['dur'],entrega=str(x['entrega']),
                       posteriores_concluidos_antes=later_done))
    return dict(limite_iqr=round(float(hi),0),n_outliers=len(outs),outliers=ex,
                aguardando_extremos=[dict(nome=x['nome'],dias=x['dur'],entrega=str(x['entrega'])) for x in wait])
out['outliers']={pid:outliers(pid) for pid in DATA}

# ---------------------------------------------------------------
# 13/14. SIMULAÇÃO via KM (Monte Carlo do tempo restante)
# ---------------------------------------------------------------
def simulate(pid, entrega_iso):
    r=DATA[pid]
    km,_=kaplan_meier(r)
    ts=[x[0] for x in km]; Ss=[x[1] for x in km]
    entrega=to_date(entrega_iso); elapsed=(TODAY-entrega).days
    def S(t):
        s=[Ss[i] for i in range(len(ts)) if ts[i]<=t]
        return s[-1] if s else 1.0
    S0=S(elapsed) or 1e-6
    S_plateau=Ss[-1] if Ss else 0.0
    # prob máxima de conclusão observável dentro do horizonte dos dados
    p_max_obs=round(1-min(1.0,S_plateau/S0),2)
    def Pcond(extra):
        return round(1-min(1.0,S(elapsed+extra)/S0),2)
    # residual condicional: mediana só se atingível
    def q_resid(qq):
        for t in ts:
            if t>=elapsed and (1-min(1.0,S(t)/S0))>=qq:
                return t-elapsed
        return None
    return dict(elapsed=elapsed,S_agora=round(S0,2),teto_conclusao_observavel=p_max_obs,
                P_mais_90d=Pcond(90),P_mais_180d=Pcond(180),P_mais_365d=Pcond(365),
                residual_p25=q_resid(0.25),residual_p50=q_resid(0.50),residual_p75=q_resid(0.75))
# exemplos de simulação
sim_examples={}
try:
    sim_examples['B_entrada_2025-08-11']=simulate('B','2025-08-11')
    sim_examples['JSI_entrada_2026-01-01']=simulate('JSI','2026-01-01')
    sim_examples['P_entrada_2026-03-01']=simulate('P','2026-03-01')
except Exception as e:
    sim_examples['erro']=str(e)
out['simulacao']=sim_examples

# throughput recente (para backlog)
def throughput(pid,months=3):
    r=[x for x in DATA[pid] if x['done']]
    cut=TODAY-dt.timedelta(days=30*months)
    rec=[x for x in r if x['resol']>=cut]
    return round(len(rec)/months,1)
out['throughput_mensal_3m']={pid:throughput(pid) for pid in DATA}

# ---------------------------------------------------------------
# 7. BACKLOG OCULTO (triangulação com fração de captura f)
# ---------------------------------------------------------------
def backlog(pid):
    r=DATA[pid]
    waiting=sum(1 for x in r if not x['done'])
    thr=throughput(pid,3)
    res=dict(observado_aguardando=waiting,throughput_mensal=thr)
    # cenários de fração de captura (quanto da fila real aparece na planilha)
    for f in (0.7,0.5,0.3):
        real=round(waiting/f)
        res[f'f_{int(f*100)}pct']=dict(fila_real_estimada=real,nao_observados=real-waiting,
            meses_p_zerar=round(real/thr,1) if thr>0 else None)
    return res
out['backlog']={pid:backlog(pid) for pid in DATA}

print(json.dumps(out,ensure_ascii=False,default=str,indent=1))
# salva
with open(r'C:\temp\Workshop Cursor\consulado\scripts\insights.json','w',encoding='utf-8') as f:
    json.dump(out,f,ensure_ascii=False,default=str,indent=1)
