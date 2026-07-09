#!/usr/bin/env python3
"""
Fetch sheets from OneDrive and generate dados.json
Designed to run via GitHub Actions
"""

import json
import sys
import re
from datetime import datetime
from io import BytesIO
from urllib.request import urlopen, Request
from urllib.error import URLError

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)


# URLs das planilhas do OneDrive - com parâmetro download=1
SHEETS = {
    'B': 'https://1drv.ms/x/c/ba23482b38fbdc1e/EVyxoIPnz65Du-mLBAl7CuUBOsLybUkq3f-TJiOIyW9HBA?download=1',
    'JSI': 'https://1drv.ms/x/c/ba23482b38fbdc1e/ETkw9cQ64L5Ko2WicjOlg04BiOc1mm4MLMiqKWcidPm6iA?download=1',
    'P': 'https://1drv.ms/x/c/ba23482b38fbdc1e/EekjF_2Qd5tJoDqaLQ8-IokB436GSnReRCEBHgeuyg72uA?download=1',
}


def normalize_date(val):
    """Convert Excel date to ISO 8601 string (YYYY-MM-DD)"""
    if val is None or val == '':
        return None
    
    if isinstance(val, str):
        # Try common date formats
        val = val.strip()
        for fmt in ['%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                dt = datetime.strptime(val, fmt)
                return dt.strftime('%Y-%m-%d')
            except (ValueError, TypeError):
                pass
        return None
    
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    
    # Excel numeric date (days since 1900-01-01)
    if isinstance(val, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            dt = excel_epoch + __import__('datetime').timedelta(days=val)
            return dt.strftime('%Y-%m-%d')
        except:
            return None
    
    return None


def download_sheet(url, sheet_name):
    """Download Excel file from OneDrive"""
    try:
        print(f"  ⏳ Baixando planilha {sheet_name}...")
        
        # Cria request com User-Agent para evitar bloqueios
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        req = Request(url, headers=headers)
        
        with urlopen(req, timeout=30) as response:
            content = response.read()
            if not content:
                print(f"  ❌ Resposta vazia do OneDrive")
                return None
            return BytesIO(content)
    
    except URLError as e:
        print(f"  ❌ Erro de rede: {e}")
        return None
    except Exception as e:
        print(f"  ❌ Erro ao baixar: {type(e).__name__}: {e}")
        return None


def parse_excel_sheet(buffer, sheet_name):
    """Parse Excel worksheet and extract records"""
    try:
        # Lê planilha com data_only=True para valores em vez de fórmulas
        wb = openpyxl.load_workbook(buffer, data_only=True)
        
        # Tenta encontrar sheet com nome exato, senão usa primeira ativa
        ws = None
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Tenta encontrar por padrão
            for sn in wb.sheetnames:
                if sheet_name.lower() in sn.lower() or sn.lower() in sheet_name.lower():
                    ws = wb[sn]
                    break
        
        if not ws:
            ws = wb.active
        
        records = []
        skip_patterns = ['MEDIA', 'MÉDIA', 'MAIOR', 'MENOR', 'TOTAL', 'RECEBIDO', 'ATUALIZ', 'N°', 'Nº', 'PLANILHA', 'ESTAMOS']
        
        # Itera a partir da linha 2 (pulando header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(cell is not None for cell in row):
                continue
            
            # Esperado: [Nº, Nome, Sigla, Data Entrega, Resolução, Grupo, ...]
            # Mas ajustamos para ser flexível
            col_vals = [cell for cell in row]
            
            # Tenta extrair com base em posição (ajusta conforme a estrutura real)
            nome = col_vals[1] if len(col_vals) > 1 else None
            sigla = col_vals[2] if len(col_vals) > 2 else None
            entrega = col_vals[3] if len(col_vals) > 3 else None
            resol = col_vals[4] if len(col_vals) > 4 else None
            grupo = col_vals[5] if len(col_vals) > 5 else None
            
            # Se nome for nulo, tenta inverter (às vezes sigla está em nome)
            if not nome and sigla:
                nome = sigla
                sigla = None
            
            # Converte para string e limpa
            nome_str = str(nome).strip() if nome else ''
            sigla_str = str(sigla).strip() if sigla else ''
            
            # Pula se for linha de cabeçalho ou agregação
            if not nome_str:
                continue
            if any(p.upper() in nome_str.upper() for p in skip_patterns):
                continue
            
            # Normaliza datas
            entrega_iso = normalize_date(entrega)
            resol_iso = normalize_date(resol) if resol else None
            
            # Só inclui se tem data de entrega
            if not entrega_iso:
                continue
            
            # Cria registro
            record = {
                'entrega': entrega_iso,
                'nome': (sigla_str or nome_str)[:50],  # Limita tamanho
                'grupo': str(grupo).strip()[:10] if grupo else '1',
                'resol': resol_iso,
                'planilha': sheet_name,
            }
            
            records.append(record)
        
        print(f"  ✅ {len(records)} registros processados")
        return records
    
    except Exception as e:
        print(f"  ❌ Erro ao processar: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return []


def consolidate_data():
    """Download all sheets and consolidate"""
    all_records = []
    
    print("📥 Iniciando download das planilhas...\n")
    
    for sheet_name, url in SHEETS.items():
        print(f"📊 Planilha {sheet_name}:")
        buffer = download_sheet(url, sheet_name)
        if buffer:
            try:
                records = parse_excel_sheet(buffer, sheet_name)
                all_records.extend(records)
            except Exception as e:
                print(f"  ❌ Erro fatal: {e}")
        else:
            print(f"  ⚠️  Não foi possível baixar")
    
    # Remove duplicatas (mantém primeira ocorrência)
    seen = set()
    unique_records = []
    for r in all_records:
        key = (r['entrega'], r['nome'])
        if key not in seen:
            seen.add(key)
            unique_records.append(r)
    
    print(f"\n📈 Total consolidado: {len(unique_records)} registros únicos")
    return unique_records


def save_json(records, filename='dados.json'):
    """Save consolidated data to JSON"""
    payload = {
        'meta': {
            'generated_at': datetime.utcnow().isoformat() + 'Z',
            'source': 'GitHub Actions + OneDrive',
            'count': len(records),
        },
        'records': records,
    }
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ Salvo: {filename} ({len(records)} registros)")
        return True
    except Exception as e:
        print(f"\n❌ Erro ao salvar: {e}")
        return False


if __name__ == '__main__':
    try:
        records = consolidate_data()
        if records:
            success = save_json(records)
            sys.exit(0 if success else 1)
        else:
            print("\n⚠️  Nenhum registro foi lido. Criando dados.json vazio...")
            save_json([])
            sys.exit(0)
    except Exception as e:
        print(f"\n❌ Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
