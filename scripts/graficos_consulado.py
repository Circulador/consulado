# -*- coding: utf-8 -*-
"""Gera figuras da engenharia reversa (KM, hazard, distribuicoes, familia)."""
import os, re, math, datetime as dt
import numpy as np
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import gridspec
from sklearn.mixture import GaussianMixture

TODAY = dt.date(2026,7,17)
OUT = r'C:\temp\Workshop Cursor\consulado\scripts'
FILES = {
 'B':  r'C:\Users\81008280\Downloads\B APOS ENTREGA NO BALCAO ATE EMAIL  JA SOU ITALIANO (10).xlsx',
 'JSI':r'C:\Users\81008280\Downloads\JSI apos receber e mail Ja sou Italiano aguardando finalizaco consular RJ (1).xlsx',
 'P':  r'C:\Users\81008280\Downloads\P Espera por Pendencias Tempo Consulado SEDEX (1).xlsx',
}
LAYOUT={'B':(0,1,2,4),'JSI':(0,1,2,4),'P':(1,0,2,4)}
BAD=re.compile(r'^(MEDIA|M[EÉ]DIA|MENOR|MAIOR|TOTAL|ULTIMO|[UÚ]LTIMO|ESTAMOS|DATA DE|NOME|RECEBIDO|GRUPO|N[ºo°]|B PLAN|PLANILHA|ATUALIZAD|DIAS|NAN)',re.I)
def to_date(v):
    if isinstance(v,dt.datetime): return v.date()
    if isinstance(v,dt.date): return v
    return None
def load(pid):
    ce,cn,cg,cr=LAYOUT[pid]
    wb=openpyxl.load_workbook(FILES[pid],data_only=True,read_only=True)
    ws=wb[wb.sheetnames[0]]; recs=[]
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i<2 or row is None or len(row)<=max(ce,cn,cr): continue
        nome=row[cn]; nome=str(nome).strip() if nome is not None else ''
        ent=to_date(row[ce])
        if not nome or BAD.match(nome) or ent is None: continue
        res=to_date(row[cr])
        d=(res-ent).days if res else (TODAY-ent).days
        if d<0: continue
        recs.append(dict(planilha=pid,nome=re.sub(r'\s+',' ',nome),entrega=ent,resol=res,done=res is not None,dur=d))
    return recs
DATA={pid:load(pid) for pid in FILES}

COL={'B':'#2a78d6','JSI':'#1baf7a','P':'#eda100'}
plt.rcParams.update({'font.size':10,'axes.grid':True,'grid.alpha':.25,'figure.dpi':130})

def km(recs):
    arr=sorted([(x['dur'],1 if x['done'] else 0) for x in recs])
    times=sorted(set(t for t,_ in arr)); S=1.0; xs=[0];ys=[1.0]
    for t in times:
        d=sum(1 for tt,e in arr if tt==t and e==1); r=sum(1 for tt,e in arr if tt>=t)
        if r>0 and d>0: S*=(1-d/r)
        xs.append(t); ys.append(S)
    return xs,ys

# ---- FIG 1: Kaplan-Meier 3 filas ----
fig,ax=plt.subplots(figsize=(9,5))
for pid in ['B','JSI','P']:
    xs,ys=km(DATA[pid])
    ax.step(xs,ys,where='post',lw=2.2,color=COL[pid],
            label=f"{pid} (n={len(DATA[pid])}, {sum(x['done'] for x in DATA[pid])} eventos)")
ax.axhline(.5,ls='--',c='#888',lw=1); ax.set_ylim(0,1)
ax.set_xlim(0,560); ax.set_xlabel('Dias na fila'); ax.set_ylabel('S(t) = prob. de AINDA estar na fila')
ax.set_title('Curvas de Sobrevivência (Kaplan-Meier) — 3 filas do consulado')
ax.legend(loc='upper right')
ax.annotate('Platô de B em ~0,36:\n~36% não conclui no\nhorizonte observável',
            xy=(430,.36),xytext=(300,.62),fontsize=9,color='#b91c1c',
            arrowprops=dict(arrowstyle='->',color='#b91c1c'))
fig.tight_layout(); fig.savefig(os.path.join(OUT,'fig1_km.png')); plt.close(fig)

# ---- FIG 2: Hazard por bin de 30d (B e JSI) ----
def hazard(recs,bw=30,tmax=360):
    arr=[(x['dur'],1 if x['done'] else 0) for x in recs]; rows=[]
    for lo in range(0,tmax,bw):
        hi=lo+bw
        r=sum(1 for t,e in arr if t>=lo); w=sum(1 for t,e in arr if lo<=t<hi and e==0)
        d=sum(1 for t,e in arr if lo<=t<hi and e==1); eff=r-w/2
        rows.append((lo+bw/2, d/eff if eff>0 else 0))
    return zip(*rows)
fig,ax=plt.subplots(figsize=(9,5))
for pid in ['B','JSI']:
    xs,hs=hazard(DATA[pid])
    ax.plot(list(xs),list(hs),'o-',lw=2,color=COL[pid],label=pid)
ax.set_xlabel('Dias na fila'); ax.set_ylabel('Hazard (prob. de sair no intervalo)')
ax.set_title('Taxa de Risco (Hazard) — janelas de processamento')
ax.axvspan(270,300,color='#2a78d6',alpha=.12)
ax.annotate('Janela de B\n~270–300 dias\n(hazard 0,38)',xy=(285,.38),xytext=(150,.33),
            fontsize=9,color='#1e4e9e',arrowprops=dict(arrowstyle='->',color='#1e4e9e'))
ax.annotate('Pico de JSI\n~30–60 e 210–240 d',xy=(45,.152),xytext=(60,.28),
            fontsize=9,color='#0d7a52',arrowprops=dict(arrowstyle='->',color='#0d7a52'))
ax.legend(); fig.tight_layout(); fig.savefig(os.path.join(OUT,'fig2_hazard.png')); plt.close(fig)

# ---- FIG 3: Distribuicao + GMM (B e JSI) ----
fig=plt.figure(figsize=(11,4.5)); gs=gridspec.GridSpec(1,2)
for k,pid in enumerate(['B','JSI']):
    ax=fig.add_subplot(gs[k])
    d=np.array([x['dur'] for x in DATA[pid] if x['done'] and x['dur']>0],float)
    ax.hist(d,bins=20,color=COL[pid],alpha=.45,density=True,edgecolor='white')
    X=np.log(d).reshape(-1,1)
    kbest=3 if pid=='JSI' else 4
    g=GaussianMixture(kbest,n_init=5,random_state=0).fit(X)
    xs=np.linspace(1,d.max(),400); lx=np.log(xs).reshape(-1,1)
    dens=np.exp(g.score_samples(lx))/xs
    ax.plot(xs,dens,'k-',lw=2,label=f'GMM {kbest} comp.')
    for i in range(kbest):
        c=np.exp(g.means_[i,0]); ax.axvline(c,ls=':',c='#c0392b',alpha=.6)
    ax.set_title(f'{pid}: distribuição dos tempos concluídos (n={len(d)})')
    ax.set_xlabel('Dias'); ax.set_ylabel('densidade'); ax.legend()
fig.suptitle('Distribuições multimodais → múltiplos regimes internos',y=1.02)
fig.tight_layout(); fig.savefig(os.path.join(OUT,'fig3_dist.png'),bbox_inches='tight'); plt.close(fig)

# ---- FIG 4: Pasta familiar (boxplot familia vs solo) ----
def split_family(pid):
    r=[x for x in DATA[pid] if x['done'] and x['dur']>0]
    suf={}
    for x in r:
        p=re.split(r'[-–/]',x['nome'])
        if len(p)>=2:
            s=p[-1].strip().upper()
            if len(s)>=2: suf.setdefault(s,[]).append(x)
    fam=set()
    for s,v in suf.items():
        if len(v)>1:
            for x in v: fam.add(id(x))
    for x in r:
        if re.search(r'IRM[ÃA]O?S?',x['nome'],re.I): fam.add(id(x))
    return ([x['dur'] for x in r if id(x) in fam],[x['dur'] for x in r if id(x) not in fam])
fig,axes=plt.subplots(1,2,figsize=(10,4.5))
for ax,pid in zip(axes,['B','JSI']):
    f,s=split_family(pid)
    bp=ax.boxplot([f,s],tick_labels=[f'Família\n(proxy) n={len(f)}',f'Solo\nn={len(s)}'],
                  patch_artist=True,widths=.5)
    for patch,c in zip(bp['boxes'],['#1baf7a','#c0392b']): patch.set_facecolor(c); patch.set_alpha(.5)
    ax.set_title(f'{pid}: tempo por tipo (proxy pasta familiar)')
    ax.set_ylabel('Dias até conclusão')
fig.suptitle('Hipótese Pasta Familiar: grupos com sobrenome compartilhado são mais rápidos',y=1.02)
fig.tight_layout(); fig.savefig(os.path.join(OUT,'fig4_familia.png'),bbox_inches='tight'); plt.close(fig)

print('figuras geradas:', [f for f in os.listdir(OUT) if f.endswith('.png')])
